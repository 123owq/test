"""
Tool 8: Excel Claim Chart 생성기
파이프라인 결과 → compareFormt.xlsx 형식의 Excel 파일 출력
저장 경로: data/results/{출원번호}/round_{차수}_chart.xlsx
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from schemas.chart import ClaimChart, DiffAnalysis, AmendedClaim
from schemas.office_action import OfficeAction

# ── 색상 상수 ────────────────────────────────────────────────────────────────
_FILL_HEADER      = PatternFill("solid", fgColor="1F3864")  # 진한 남색 — 표 헤더
_FILL_SECTION     = PatternFill("solid", fgColor="D6DCE4")  # 연회색 — 섹션 타이틀
_FILL_IDENTICAL   = PatternFill("solid", fgColor="FFFF00")  # 노란색 — 동일/균등
_FILL_PARTIAL     = PatternFill("solid", fgColor="FFD966")  # 주황 — 부분 차이
_FILL_NOT_FOUND   = PatternFill("solid", fgColor="92D050")  # 초록 — 미개시(핵심 차이)
_FILL_CLAIM_ROW   = PatternFill("solid", fgColor="F2F2F2")  # 연회색 — 당사 청구항 열

_FONT_TITLE  = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
_FONT_HEADER = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
_FONT_BODY   = Font(name="맑은 고딕", size=9)
_FONT_BOLD   = Font(name="맑은 고딕", bold=True, size=9)
_FONT_SECTION= Font(name="맑은 고딕", bold=True, size=10)

_THIN  = Side(style="thin",   color="BFBFBF")
_THICK = Side(style="medium", color="4472C4")
_BORDER_ALL   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BORDER_THICK = Border(left=_THICK, right=_THICK, top=_THICK, bottom=_THICK)

_BASE_DIR = Path(__file__).parent.parent / "data" / "results"


# ── 헬퍼 ─────────────────────────────────────────────────────────────────────

def _cell(ws, row: int, col: int, value=None, font=None, fill=None,
          align=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font      = font
    if fill:   c.fill      = fill
    if align:  c.alignment = align
    if border: c.border    = border
    return c


def _merge(ws, r1, c1, r2, c2, value=None, font=None, fill=None, align=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    if font:  c.font      = font
    if fill:  c.fill      = fill
    if align: c.alignment = align
    return c


def _wrap_center(bold=False):
    return Alignment(wrap_text=True, vertical="center", horizontal="center")


def _wrap_left():
    return Alignment(wrap_text=True, vertical="top", horizontal="left")


def _similarity_fill(similarity: str) -> PatternFill:
    return {
        "identical":           _FILL_IDENTICAL,
        "equivalent":          _FILL_IDENTICAL,
        "partially_different": _FILL_PARTIAL,
        "not_found":           _FILL_NOT_FOUND,
    }.get(similarity, PatternFill())


def _similarity_label(similarity: str) -> str:
    return {
        "identical":           "동일",
        "equivalent":          "균등",
        "partially_different": "부분 차이",
        "not_found":           "미개시",
    }.get(similarity, similarity)


# ── 메인 함수 ─────────────────────────────────────────────────────────────────

def export_excel(
    office_action: OfficeAction,
    charts: list[ClaimChart],
    diff: DiffAnalysis,
    amended: AmendedClaim,
    round_no: int,
) -> str:
    """
    파이프라인 결과를 compareFormt.xlsx 형식의 Excel로 저장.
    반환: 저장된 파일 경로 문자열
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Claim Chart"

    # 인용발명 목록 (charts 순서대로)
    prior_arts = [c.prior_art_id for c in charts]
    n_pa = len(prior_arts)          # 인용발명 개수
    total_cols = 2 + n_pa + 1       # 청구항(1) + 구성요소(2) + 인용발명들 + 비고

    # 컬럼 인덱스 편의 상수
    COL_NO    = 1   # 구성요소 번호
    COL_CLAIM = 2   # 당사 청구항
    COL_PA    = [3 + i for i in range(n_pa)]  # 인용발명 열들
    COL_NOTE  = 3 + n_pa                       # 비고

    # ── 컬럼 너비 설정 ──────────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(COL_NO)].width    = 8
    ws.column_dimensions[get_column_letter(COL_CLAIM)].width = 45
    for col in COL_PA:
        ws.column_dimensions[get_column_letter(col)].width   = 45
    ws.column_dimensions[get_column_letter(COL_NOTE)].width  = 25

    row = 1  # 현재 행 포인터

    # ════════════════════════════════════════════════════════════════════════
    # 1. 사건번호 헤더
    # ════════════════════════════════════════════════════════════════════════
    _merge(ws, row, 1, row, total_cols,
           value=f"사건번호: {office_action.application_no}　　발명의 명칭: {office_action.title}",
           font=_FONT_TITLE, fill=_FILL_HEADER,
           align=_wrap_center())
    ws.row_dimensions[row].height = 22
    row += 1

    # 체크박스 행 (텍스트로 표현)
    _merge(ws, row, 1, row, total_cols,
           value="□ 수정없음(원본)          □ 의견제출(반박)          □ 기재불비(보정)",
           font=_FONT_BODY, align=_wrap_center())
    ws.row_dimensions[row].height = 18
    row += 1

    # ════════════════════════════════════════════════════════════════════════
    # 2. 거절 부분
    # ════════════════════════════════════════════════════════════════════════
    _merge(ws, row, 1, row, total_cols,
           value="2. 거절 부분",
           font=_FONT_SECTION, fill=_FILL_SECTION, align=_wrap_left())
    ws.row_dimensions[row].height = 18
    row += 1

    rejection_summary = "　".join(
        f"[거절이유 {r.seq}] {r.summary}" for r in office_action.rejection_reasons
    )
    _merge(ws, row, 1, row, total_cols,
           value=f"□ 당사 거절    □ 출원인 거절 기준 심사    □ 기재불비(보정) 검토\n{rejection_summary}",
           font=_FONT_BODY, align=_wrap_left())
    ws.row_dimensions[row].height = 36
    row += 1

    # ════════════════════════════════════════════════════════════════════════
    # 3. Claim Chart 표
    # ════════════════════════════════════════════════════════════════════════
    _merge(ws, row, 1, row, total_cols,
           value="3. 구성요소별 대비표 (Claim Chart)",
           font=_FONT_SECTION, fill=_FILL_SECTION, align=_wrap_left())
    ws.row_dimensions[row].height = 18
    row += 1

    # 표 헤더 행
    _cell(ws, row, COL_NO,    "구성",    _FONT_HEADER, _FILL_HEADER, _wrap_center(), _BORDER_ALL)
    _cell(ws, row, COL_CLAIM, "청구항",  _FONT_HEADER, _FILL_HEADER, _wrap_center(), _BORDER_ALL)
    for i, pa_id in enumerate(prior_arts):
        pa_ref = next(
            (pa.reference for pa in office_action.cited_prior_arts if pa.id == pa_id),
            pa_id
        )
        _cell(ws, row, COL_PA[i],
              f"{pa_id}\n({pa_ref})",
              _FONT_HEADER, _FILL_HEADER, _wrap_center(), _BORDER_ALL)
    _cell(ws, row, COL_NOTE, "비고", _FONT_HEADER, _FILL_HEADER, _wrap_center(), _BORDER_ALL)
    ws.row_dimensions[row].height = 30
    row += 1

    # 도면 행 (placeholder)
    _cell(ws, row, COL_NO,    "도면", _FONT_BOLD, _FILL_CLAIM_ROW, _wrap_center(), _BORDER_ALL)
    _cell(ws, row, COL_CLAIM, "(도면 첨부)", _FONT_BODY, None, _wrap_center(), _BORDER_ALL)
    for col in COL_PA:
        _cell(ws, row, col, "(인용발명 도면)", _FONT_BODY, None, _wrap_center(), _BORDER_ALL)
    _cell(ws, row, COL_NOTE, "", _FONT_BODY, None, _wrap_center(), _BORDER_ALL)
    ws.row_dimensions[row].height = 60
    row += 1

    # 구성요소별 행 — 첫 번째 chart(대표 독립항) 기준으로 행 구성
    # 각 구성요소에 대해 인용발명별 mapping을 매칭
    if charts:
        base_chart = charts[0]
        for mapping in base_chart.mappings:
            elem_id   = mapping.our_element.element_id
            elem_text = mapping.our_element.text

            _cell(ws, row, COL_NO,    elem_id,   _FONT_BOLD, _FILL_CLAIM_ROW, _wrap_center(), _BORDER_ALL)
            _cell(ws, row, COL_CLAIM, elem_text, _FONT_BODY, _FILL_CLAIM_ROW, _wrap_left(),   _BORDER_ALL)

            for i, chart in enumerate(charts):
                # 같은 element_id의 mapping 찾기
                m = next(
                    (m for m in chart.mappings if m.our_element.element_id == elem_id),
                    mapping  # 없으면 기본 mapping 사용
                )
                fill = _similarity_fill(m.similarity)
                label = _similarity_label(m.similarity)
                cell_text = f"[{label}]\n{m.prior_art_text}\n\n{m.analysis}"
                _cell(ws, row, COL_PA[i], cell_text, _FONT_BODY, fill, _wrap_left(), _BORDER_ALL)

            _cell(ws, row, COL_NOTE, "", _FONT_BODY, None, _wrap_left(), _BORDER_ALL)
            ws.row_dimensions[row].height = 80
            row += 1

    # ════════════════════════════════════════════════════════════════════════
    # 4. 의견 요지 및 구성요소 분석
    # ════════════════════════════════════════════════════════════════════════
    row += 1
    _merge(ws, row, 1, row, total_cols,
           value="4. 의견 요지 및 구성요소 분석",
           font=_FONT_SECTION, fill=_FILL_SECTION, align=_wrap_left())
    ws.row_dimensions[row].height = 18
    row += 1

    # 핵심 차이점
    diff_text = "\n".join(f"• {d}" for d in diff.key_differences)
    _merge(ws, row, 1, row, total_cols,
           value=f"[핵심 차이점]\n{diff_text}",
           font=_FONT_BODY, align=_wrap_left())
    ws.row_dimensions[row].height = max(18 * len(diff.key_differences), 40)
    row += 1

    # 거절이유별 전략
    for s in diff.strategies:
        strategy_text = (
            f"[거절이유 {s.rejection_seq}] "
            f"전략: {s.strategy_type}　　{s.rationale}"
        )
        _merge(ws, row, 1, row, total_cols,
               value=strategy_text, font=_FONT_BODY, align=_wrap_left())
        ws.row_dimensions[row].height = 36
        row += 1

    # ════════════════════════════════════════════════════════════════════════
    # 5. 보정 내용
    # ════════════════════════════════════════════════════════════════════════
    row += 1
    _merge(ws, row, 1, row, total_cols,
           value="5. 보정 내용",
           font=_FONT_SECTION, fill=_FILL_SECTION, align=_wrap_left())
    ws.row_dimensions[row].height = 18
    row += 1

    # 원본 vs 보정 2열 레이아웃
    half = total_cols // 2
    _merge(ws, row, 1, row, half,
           value=f"[원본 청구항 제{amended.original_claim_number}항]",
           font=_FONT_BOLD, fill=_FILL_SECTION, align=_wrap_center())
    _merge(ws, row, half + 1, row, total_cols,
           value="[보정 청구항]",
           font=_FONT_BOLD, fill=_FILL_SECTION, align=_wrap_center())
    ws.row_dimensions[row].height = 18
    row += 1

    orig_lines  = amended.original_text.count("\n") + 1
    amend_lines = amended.amended_text.count("\n") + 1
    row_h = max(orig_lines, amend_lines) * 15 + 10

    _merge(ws, row, 1, row, half,
           value=amended.original_text,
           font=_FONT_BODY, align=_wrap_left())
    _merge(ws, row, half + 1, row, total_cols,
           value=amended.amended_text,
           font=_FONT_BODY, align=_wrap_left())
    ws.row_dimensions[row].height = row_h
    row += 1

    # 보정 이유
    _merge(ws, row, 1, row, total_cols,
           value=f"[보정 이유]\n{amended.amendment_rationale}",
           font=_FONT_BODY, align=_wrap_left())
    ws.row_dimensions[row].height = max(amended.amendment_rationale.count("\n") * 15 + 20, 60)
    row += 1

    # 해소된 거절이유
    _merge(ws, row, 1, row, total_cols,
           value=f"해소된 거절이유: {amended.addresses_rejections}　　품질 점수: {amended.quality_score}",
           font=_FONT_BODY, align=_wrap_left())
    ws.row_dimensions[row].height = 18

    # ── 저장 ──────────────────────────────────────────────────────────────
    out_dir = _BASE_DIR / office_action.application_no
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"round_{round_no:02d}_chart.xlsx"
    wb.save(str(out_path))
    return str(out_path)
